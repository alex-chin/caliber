# -*- coding: utf-8 -*-
"""
Created on Sat Apr 28 15:30:37 2018

@author: Alex
"""
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from math import nan


# Запись прайса и форматирование
class Write_price:
    def __init__(self, name, datafr):
        self.name = name
        self.data = datafr
        self.desc_fields = []

    def open_wb(self, header):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.header = header

    def _get_field_descs(self, dops):
        fields = ['catroom', 'meal', 'period', 'sgl', 'dbl', 'trp', 'age', 'four','dsu']
        descs = ['', '', '', '1', '2', '3', '','4', 'DSU']
        for dop in dops:
            fields.append(dop.name)
            descs.append(dop.desc)
        # упорядочивание полей
        self.data = self.data[fields]
        return descs

    def write(self, dop_name):
        descs = self._get_field_descs(dop_name)
        self.ws.append(descs)
        self.ws.append([''])
        self.ws.append([self.header])

        begin_col = 1
        finish_col = begin_col + len(descs) - 1
        self.ws.merge_cells(start_row=2, start_column=begin_col, end_row=2, end_column=finish_col)
        self.ws.merge_cells(start_row=3, start_column=begin_col, end_row=3, end_column=finish_col)

        for r in dataframe_to_rows(self.data, index=False, header=False):
            self.ws.append(r)

        for cell in self.ws['A'] + self.ws['B'] + self.ws['C']:
            cell.alignment = Alignment(horizontal='left')

        for cell in self.ws['d'] + self.ws['e'] + self.ws['f'] + self.ws['g'] + self.ws['h']:
            cell.alignment = Alignment(horizontal='right')

        self.ws['a3'].alignment = Alignment(horizontal='center')

        self.ws.column_dimensions['a'].width = 22
        self.ws.column_dimensions['b'].width = 20
        self.ws.column_dimensions['c'].width = 22
        self.ws.row_dimensions[1].height = 25

    def close(self):
        self.wb.save(self.name)
        self.wb.close()
